#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""逐行对比 派单.xlsx 与 AI通报.xlsx，校验新规则下的数据准确性"""

import openpyxl
from collections import defaultdict
import os
import sys

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STAFF_MAP = {
    "数字政务": 11, "交通物流": 7, "教育": 6,
    "开发区政商": 10, "荔城政商": 9, "数字企业": 7, "新塘政商": 11
}


def get_col_idx(ws, header_name, row_num=3):
    """在指定行查找包含 header_name 的列索引（支持模糊匹配）"""
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row_num, col).value
        if val:
            val_str = str(val).strip()
            if header_name in val_str:
                return col
    return None


def calc_from_raw(records):
    """从原始记录计算基础指标"""
    total = len(records)
    collab = sum(1 for r in records if r["collab"])
    dir_cnt = sum(1 for r in records if "总监" in r["collab"])
    ya_cnt = sum(1 for r in records if "云安" in r["collab"])

    opp = [r for r in records if r["is_opp"] == "是"]
    proj = sum(1 for r in opp if r["is_project"] == "是")
    std = len(opp) - proj
    amt = sum(r["amount"] for r in opp)

    conv = [r for r in records if r["is_convert"] == "是"]
    conv_cnt = len(conv)
    conv_amt = sum(r["amount"] for r in conv)
    zero = len(set(r["manager"] for r in conv if r["manager"]))

    new_ent = [r for r in records if r["is_new_ent"] == "是"]
    new_cnt = len(new_ent)
    new_conv = sum(1 for r in new_ent if r["is_convert"] == "是")
    new_rate = new_conv / new_cnt if new_cnt else 0

    sc = {"养虾": 0, "智能体": 0, "知识库": 0, "远航平台": 0}
    for r in conv:
        s = r["scenario"]
        for k in sc:
            if k in s: sc[k] += 1; break

    return {
        "走访数": total, "协同走访数": collab, "协同走访占比": collab / total if total else 0,
        "其中：总监": dir_cnt, "其中：云安": ya_cnt, "标品商机数": std, "项目商机数": proj,
        "总商机金额": amt, "转化数": conv_cnt, "转化金额": conv_amt,
        "破零人数": zero, "新增商企数": new_cnt, "新增商企 AI 渗透率": new_rate,
        **sc
    }


def main():
    print("正在加载数据...")
    # 1. 读派单
    wb1 = openpyxl.load_workbook(os.path.join(BASE_DIR, "派单.xlsx"), data_only=True)
    ws1 = wb1.worksheets[0]
    records = []
    for r in range(2, ws1.max_row + 1):
        dept = (ws1.cell(r, 1).value or "").strip()
        if not dept: continue
        records.append({
            "dept": dept, "manager": (ws1.cell(r, 3).value or "").strip(),
            "collab": (ws1.cell(r, 5).value or "").strip(),
            "is_opp": (ws1.cell(r, 6).value or "").strip(),
            "amount": float(ws1.cell(r, 8).value or 0),
            "is_convert": (ws1.cell(r, 11).value or "").strip(),
            "is_project": (ws1.cell(r, 12).value or "").strip(),
            "is_new_ent": (ws1.cell(r, 13).value or "").strip(),
            "scenario": (ws1.cell(r, 15).value or "").strip()
        })
    wb1.close()

    groups = defaultdict(list)
    for rec in records: groups[rec["dept"]].append(rec)
    manual = {d: calc_from_raw(g) for d, g in groups.items()}

    # 2. 读通报
    wb2 = openpyxl.load_workbook(os.path.join(BASE_DIR, "AI通报.xlsx"), data_only=True)
    ws2 = wb2["通报模板"]

    keys = ["走访数", "协同走访数", "协同走访占比", "其中：总监", "其中：云安",
            "标品商机数", "项目商机数", "总商机金额", "转化数", "转化金额",
            "破零人数", "AI 破零率", "新增商企数", "新增商企 AI 渗透率",
            "养虾", "智能体", "知识库", "远航平台"]

    idx_map = {k: get_col_idx(ws2, k, row_num=3) for k in keys if get_col_idx(ws2, k, row_num=3)}
    dept_col = get_col_idx(ws2, "单位", row_num=3) or 1

    # 🔧 核心修复：预先扫描找到“合计”行，严格限制循环范围，防止读到下方的脏数据
    end_row = ws2.max_row
    for r in range(4, ws2.max_row + 1):
        cell_val = ws2.cell(r, dept_col).value
        if cell_val and str(cell_val).strip() == "合计":
            end_row = r
            break

    print(f"\n️ 识别到数据范围：第 4 行 至 第 {end_row - 1} 行 (合计行在第 {end_row} 行)")

    print("\n" + "=" * 80)
    print("   数据校验报告 (v2.3 - 锁定数据范围)")
    print("=" * 80)
    errors = []

    # 3. 校验各部门数据 (只遍历到合计行之前)
    for row in range(4, end_row):
        dept = (ws2.cell(row, dept_col).value or "").strip()

        # 跳过空行
        if not dept:
            continue

        # 如果部门名不在字典里，可能是备注或其他文字，跳过
        if dept not in manual:
            print(f"⚠️ 第 {row} 行发现未知内容：'{dept}'，已跳过。")
            continue

        m = manual[dept]
        print(f"\n📦 [{dept}] (Row {row})")

        for k in keys:
            if k not in idx_map: continue
            col_idx = idx_map[k]
            act = ws2.cell(row, col_idx).value
            if act is None: act = 0

            if k == "AI 破零率":
                staff = STAFF_MAP.get(dept, 0)
                exp = m["破零人数"] / staff if staff else 0
            else:
                exp = m.get(k, 0)

            match = abs(exp - act) < 0.001 if isinstance(exp, float) else exp == act
            status = "✅" if match else "❌"
            if not match:
                errors.append(f"{dept} / {k}: 期望{exp}, 实际{act}")
            print(f"  {k:<15}  期望={str(exp):<12}  实际={str(act):<12}  {status}")

    # 4. 校验合计行
    if end_row <= ws2.max_row:
        total_manual = calc_from_raw(records)
        total_staff = sum(STAFF_MAP.values())
        exp_total_br = total_manual["破零人数"] / total_staff if total_staff else 0

        print(f"\n📊 [合计] (Row {end_row})")
        for k in keys:
            if k not in idx_map: continue
            col_idx = idx_map[k]
            act = ws2.cell(end_row, col_idx).value
            if act is None: act = 0

            if k == "AI 破零率":
                exp = exp_total_br
            else:
                exp = total_manual.get(k, 0)

            match = abs(exp - act) < 0.001 if isinstance(exp, float) else exp == act
            status = "✅" if match else "❌"
            if not match: errors.append(f"合计 / {k}: 期望{exp}, 实际{act}")
            print(f"  {k:<15}  期望={str(exp):<12}  实际={str(act):<12}  {status}")

    print("\n" + "=" * 80)
    if errors:
        print(f"⚠️ 发现 {len(errors)} 处差异:")
        for e in errors: print(f"  - {e}")
    else:
        print("🎉 全部指标校验通过！数据完全一致。")
    print("=" * 80)
    wb2.close()


if __name__ == "__main__":
    main()