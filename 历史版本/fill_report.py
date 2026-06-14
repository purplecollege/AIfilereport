#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 派单.xlsx 提取数据，按新通报规则计算指标并填入 AI通报.xlsx。

更新说明：
- 修正了数字格式：破零人数、新增商企数已恢复为整数显示
- 逻辑：破零人数按客户经理去重；AI破零率按固定编制计算
"""

import openpyxl
from openpyxl.cell.cell import MergedCell
from collections import defaultdict
import os
import sys

# ---------------------------------------------------------------------------
# 配置与常量
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PAIDAN_FILE = os.path.join(BASE_DIR, "派单.xlsx")
TONGBAO_FILE = os.path.join(BASE_DIR, "AI通报.xlsx")
SHEET_NAME = "通报模板"

# 派单.xlsx 列索引 (1-based)
COL_MAP = {
    "DEPT": 1, "MANAGER": 3, "COLLAB": 5, "IS_OPP": 6,
    "AMOUNT": 8, "IS_CONVERT": 11, "IS_PROJECT": 12,
    "IS_NEW_ENT": 13, "SCENARIO": 15
}

# 各单位固定人员总数（用于计算AI破零率）
STAFF_MAP = {
    "数字政务": 11, "交通物流": 7, "教育": 6,
    "开发区政商": 10, "荔城政商": 9, "数字企业": 7, "新塘政商": 11
}

# 通报模板输出列顺序 (A-S)
OUTPUT_COLS = [
    "走访数", "协同走访数", "协同走访占比", "其中：总监", "其中：云安",
    "标品商机数", "项目商机数", "总商机金额", "转化数", "转化金额",
    "破零人数", "AI破零率", "新增商企数", "新增商企AI渗透率",
    "养虾", "智能体", "知识库", "远航平台"
]
DATA_START_ROW = 2
OUTPUT_START_ROW = 4


# ---------------------------------------------------------------------------
# 核心函数
# ---------------------------------------------------------------------------
def read_paidan(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]
    records = []
    empty_dept = 0

    for row in range(DATA_START_ROW, ws.max_row + 1):
        dept = (ws.cell(row, COL_MAP["DEPT"]).value or "").strip()
        if not dept:
            empty_dept += 1
            continue

        manager = (ws.cell(row, COL_MAP["MANAGER"]).value or "").strip()
        collab = (ws.cell(row, COL_MAP["COLLAB"]).value or "").strip()
        is_opp = (ws.cell(row, COL_MAP["IS_OPP"]).value or "").strip()
        raw_amt = ws.cell(row, COL_MAP["AMOUNT"]).value
        is_convert = (ws.cell(row, COL_MAP["IS_CONVERT"]).value or "").strip()
        is_project = (ws.cell(row, COL_MAP["IS_PROJECT"]).value or "").strip()
        is_new_ent = (ws.cell(row, COL_MAP["IS_NEW_ENT"]).value or "").strip()
        scenario = (ws.cell(row, COL_MAP["SCENARIO"]).value or "").strip()

        try:
            amount = float(raw_amt) if raw_amt is not None and raw_amt != "" else 0.0
        except (ValueError, TypeError):
            amount = 0.0

        records.append({
            "dept": dept, "manager": manager, "collab": collab,
            "is_opp": is_opp, "amount": amount, "is_convert": is_convert,
            "is_project": is_project, "is_new_ent": is_new_ent, "scenario": scenario
        })
    wb.close()
    print(f"  ✅ 读取完成: {len(records)} 条记录 (跳过 {empty_dept} 条空部门)")
    return records


def get_dept_order(records):
    priority = ["数字政务", "数字企业", "交通物流", "教育", "新塘政商", "开发区政商", "荔城政商"]
    seen = set()
    order = []
    for r in records:
        if r["dept"] not in seen:
            order.append(r["dept"])
            seen.add(r["dept"])
    priority_map = {n: i for i, n in enumerate(priority)}
    order.sort(key=lambda d: priority_map.get(d, len(priority)))
    return order


def calc_metrics(records):
    total = len(records)
    collab_cnt = sum(1 for r in records if r["collab"])
    dir_cnt = sum(1 for r in records if "总监" in r["collab"])
    ya_cnt = sum(1 for r in records if "云安" in r["collab"])

    opp_recs = [r for r in records if r["is_opp"] == "是"]
    opp_cnt = len(opp_recs)

    proj_cnt = sum(1 for r in opp_recs if r["is_project"] == "是")
    std_cnt = opp_cnt - proj_cnt  # 标品 = 总商机 - 项目商机
    total_amt = sum(r["amount"] for r in opp_recs)

    conv_recs = [r for r in records if r["is_convert"] == "是"]
    conv_cnt = len(conv_recs)
    conv_amt = sum(r["amount"] for r in conv_recs)

    # 破零人数：去重客户经理 -> 结果为整数
    zero_breakers = set(r["manager"] for r in conv_recs if r["manager"])
    zero_breaker_cnt = len(zero_breakers)

    # 新增商企 & 渗透率 -> 结果为整数
    new_ent_recs = [r for r in records if r["is_new_ent"] == "是"]
    new_ent_cnt = len(new_ent_recs)
    new_ent_conv_cnt = sum(1 for r in new_ent_recs if r["is_convert"] == "是")
    new_ent_ai_rate = new_ent_conv_cnt / new_ent_cnt if new_ent_cnt > 0 else 0.0

    # 转化场景统计
    scenarios = {"养虾": 0, "智能体": 0, "知识库": 0, "远航平台": 0}
    for r in conv_recs:
        sc = r["scenario"]
        for key in scenarios:
            if key in sc:
                scenarios[key] += 1
                break

    return {
        "走访数": total, "协同走访数": collab_cnt,
        "协同走访占比": collab_cnt / total if total else 0,
        "其中：总监": dir_cnt, "其中：云安": ya_cnt,
        "标品商机数": std_cnt, "项目商机数": proj_cnt, "总商机金额": total_amt,
        "转化数": conv_cnt, "转化金额": conv_amt,
        "破零人数": zero_breaker_cnt,  # 整数
        "新增商企数": new_ent_cnt,  # 整数
        "新增商企AI渗透率": new_ent_ai_rate,
        **scenarios
    }


def build_summary(dept_order, dept_metrics, total_metrics):
    best_visit = max(dept_order, key=lambda d: dept_metrics[d]["走访数"])
    best_std = max(dept_order, key=lambda d: dept_metrics[d]["标品商机数"])
    best_amt = max(dept_order, key=lambda d: dept_metrics[d]["总商机金额"])

    conv_depts = [d for d in dept_order if dept_metrics[d]["转化数"] > 0]
    not_conv = [d for d in dept_order if dept_metrics[d]["转化数"] == 0]

    return "\n".join([
        f"【AI标品专项工作通报】",
        f"一、走访：累计{total_metrics['走访数']}家，协同{total_metrics['协同走访数']}({total_metrics['协同走访占比']:.0%})，{best_visit}领先({dept_metrics[best_visit]['走访数']}家)",
        f"二、商机：标品{total_metrics['标品商机数']}个，项目{total_metrics['项目商机数']}个，总金额{total_metrics['总商机金额']:,.0f}元，{best_std}标品领先，{best_amt}金额最高",
        f"三、转化：累计转化{total_metrics['转化数']}个({total_metrics['转化金额']:,.0f}元)，破零{total_metrics['破零人数']}人，{'、'.join(conv_depts)}已破零" + (
            f"，{'、'.join(not_conv)}未破零" if not_conv else "")
    ])


def safe_write(ws, row, col, value):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for mc in list(ws.merged_cells.ranges):
            if mc.min_row <= row <= mc.max_row and mc.min_col <= col <= mc.max_col:
                ws.unmerge_cells(str(mc))
                break
        cell = ws.cell(row, col)
    cell.value = value


def fill_tongbao(filepath, dept_order, dept_metrics, total_metrics, summary):
    wb = openpyxl.load_workbook(filepath)
    ws = wb[SHEET_NAME]

    total_row = OUTPUT_START_ROW + len(dept_order)
    # 清理写入区域合并单元格
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= OUTPUT_START_ROW and mc.max_row <= total_row:
            ws.unmerge_cells(str(mc))

    # 写入部门数据
    for i, dept in enumerate(dept_order):
        row = OUTPUT_START_ROW + i
        m = dept_metrics[dept]
        staff_total = STAFF_MAP.get(dept, 0)
        ai_breaker_rate = m["破零人数"] / staff_total if staff_total else 0.0

        safe_write(ws, row, 1, dept)
        for j, key in enumerate(OUTPUT_COLS):
            if key == "AI破零率":
                val = ai_breaker_rate
            else:
                val = m[key]
            safe_write(ws, row, 2 + j, val)

    # 写入合计
    safe_write(ws, total_row, 1, "合计")
    total_staff = sum(STAFF_MAP.get(d, 0) for d in dept_order)
    total_breaker_rate = total_metrics["破零人数"] / total_staff if total_staff else 0.0

    for j, key in enumerate(OUTPUT_COLS):
        if key == "AI破零率":
            val = total_breaker_rate
        else:
            val = total_metrics[key]
        safe_write(ws, total_row, 2 + j, val)

    # 🔧 关键修正：精确设置列格式
    for row in range(OUTPUT_START_ROW, total_row + 1):
        # 百分比列
        ws.cell(row, 4).number_format = "0.0%"  # 协同走访占比
        ws.cell(row, 13).number_format = "0.0%"  # AI破零率 (L列是12，这里是M列13)
        ws.cell(row, 15).number_format = "0.0%"  # 新增商企AI渗透率

        # 整数列
        ws.cell(row, 12).number_format = "0"  # 破零人数 (L列)
        ws.cell(row, 14).number_format = "0"  # 新增商企数 (N列)
        ws.cell(row, 9).number_format = "#,##0"  # 总商机金额
        ws.cell(row, 11).number_format = "#,##0"  # 转化金额

    # 写入文字摘要
    safe_write(ws, 13, 1, summary)

    wb.save(filepath)
    wb.close()
    print("  ✅ 写入完成")


def print_table(dept_order, dept_metrics, total_metrics):
    header = f"{'部门':<10} {'走访':>4} {'协同':>4} {'标品':>4} {'项目':>4} {'金额':>10} {'转化':>4} {'破零':>4} {'破零率':>6} {'新增':>4} {'渗透':>6} {'养虾':>4} {'智能':>4} {'知识':>4} {'远航':>4}"
    print("\n" + "-" * len(header))
    print(header)
    print("-" * len(header))
    for dept in dept_order:
        m = dept_metrics[dept]
        staff = STAFF_MAP.get(dept, 0)
        br = m["破零人数"] / staff if staff else 0
        # 确保控制台打印也是整数和百分比区分
        print(f"{dept:<10} {m['走访数']:>4} {m['协同走访数']:>4} {m['标品商机数']:>4} {m['项目商机数']:>4} "
              f"{m['总商机金额']:>10,.0f} {m['转化数']:>4} {m['破零人数']:>4d} {br:>5.1%} "
              f"{m['新增商企数']:>4d} {m['新增商企AI渗透率']:>5.1%} "
              f"{m['养虾']:>4} {m['智能体']:>4} {m['知识库']:>4} {m['远航平台']:>4}")
    print("-" * len(header))


def main():
    for f in [PAIDAN_FILE, TONGBAO_FILE]:
        if not os.path.exists(f):
            print(f"❌ 错误：找不到文件 {f}");
            sys.exit(1)

    print("=" * 60 + "\n   AI标品通报数据填充工具 v2.1\n" + "=" * 60)
    print("\n[1/4] 读取派单数据...")
    records = read_paidan(PAIDAN_FILE)
    if not records: print(" 无有效数据"); sys.exit(1)

    print("\n[2/4] 计算指标...")
    dept_order = get_dept_order(records)
    groups = defaultdict(list)
    for r in records: groups[r["dept"]].append(r)
    dept_metrics = {d: calc_metrics(groups[d]) for d in dept_order}
    total_metrics = calc_metrics(records)
    print_table(dept_order, dept_metrics, total_metrics)

    print("\n[3/4] 生成摘要...")
    summary = build_summary(dept_order, dept_metrics, total_metrics)
    print(f"  📝 {summary.split(chr(10))[0]}")

    print("\n[4/4] 写入通报模板...")
    fill_tongbao(TONGBAO_FILE, dept_order, dept_metrics, total_metrics, summary)

    print("\n" + "=" * 60 + "\n  ✅ 全部完成！AI通报.xlsx 已更新。\n" + "=" * 60)


if __name__ == "__main__":
    main()