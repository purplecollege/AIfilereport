#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 派单.xlsx 提取数据，按新通报规则计算指标并填入 AI通报.xlsx 的 通报模板v2 sheet。

通报模板v2 包含两个独立的表格：
  上表（4-7行数据 + 8行合计）：数字政务、数字企业、交通物流、教育
  下表（14-16行数据 + 17行合计）：新塘政商、开发区政商、荔城政商

差异：
  1. 上表第8列(H) = "项目商机数"，下表第8列(H) = "标品商机占比"
  2. N、O 两列不写入（模板中可手动删除）
  3. 标题行（第1行和第11行）动态更新为 "0506-当天日期"
"""

import openpyxl
from openpyxl.cell.cell import MergedCell
from collections import defaultdict
import os
import sys
from datetime import date

# ---------------------------------------------------------------------------
# 配置与常量
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PAIDAN_FILE = os.path.join(BASE_DIR, "派单.xlsx")
TONGBAO_FILE = os.path.join(BASE_DIR, "AI通报.xlsx")
SHEET_NAME = "通报模板v2"

# 派单.xlsx 列索引 (1-based)
COL_MAP = {
    "DEPT": 1, "MANAGER": 3, "COLLAB": 5, "IS_OPP": 6,
    "AMOUNT": 8, "OPP_TYPE": 10, "IS_CONVERT": 14, "SCENARIO": 15
}

# 各单位固定人员总数
STAFF_MAP = {
    "数字政务": 11, "交通物流": 7, "教育": 6,
    "开发区政商": 10, "荔城政商": 9, "数字企业": 7, "新塘政商": 11
}

# 上表部门（数据行4-7，合计行8）
UPPER_DEPTS = ["数字政务", "数字企业", "交通物流", "教育"]

# 下表部门（数据行14-16，合计行17），按模板指定顺序
LOWER_DEPTS = ["新塘政商", "开发区政商", "荔城政商"]

# 上表核心指标列（B-M，col 2-13），第8列=项目商机数
UPPER_CORE_KEYS = [
    "走访数", "协同走访数", "协同走访占比", "其中：总监", "其中：云安",
    "标品商机数", "项目商机数", "总商机金额", "转化数", "转化金额",
    "破零人数", "AI破零率"
]

# 下表核心指标列（B-M，col 2-13），第8列=标品商机占比
LOWER_CORE_KEYS = [
    "走访数", "协同走访数", "协同走访占比", "其中：总监", "其中：云安",
    "标品商机数", "标品商机占比", "总商机金额", "转化数", "转化金额",
    "破零人数", "AI破零率"
]

# 场景列（P-T，col 16-20），两表共用
SCENARIO_KEYS = ["养龙虾", "智能体", "知识库", "远航平台", "其他"]


# ---------------------------------------------------------------------------
# 核心函数
# ---------------------------------------------------------------------------
def read_paidan(filepath):
    """读取派单.xlsx所有记录。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]
    records = []
    empty_dept = 0

    for row in range(2, ws.max_row + 1):
        dept = (ws.cell(row, COL_MAP["DEPT"]).value or "").strip()
        if not dept:
            empty_dept += 1
            continue

        manager = (ws.cell(row, COL_MAP["MANAGER"]).value or "").strip()
        collab = (ws.cell(row, COL_MAP["COLLAB"]).value or "").strip()
        is_opp = (ws.cell(row, COL_MAP["IS_OPP"]).value or "").strip()
        raw_amt = ws.cell(row, COL_MAP["AMOUNT"]).value

        opp_type = (ws.cell(row, COL_MAP["OPP_TYPE"]).value or "").strip()
        scenario = (ws.cell(row, COL_MAP["SCENARIO"]).value or "").strip()
        is_convert = (ws.cell(row, COL_MAP["IS_CONVERT"]).value or "").strip()

        try:
            amount = float(raw_amt) if raw_amt is not None and raw_amt != "" else 0.0
        except (ValueError, TypeError):
            amount = 0.0

        records.append({
            "dept": dept, "manager": manager, "collab": collab,
            "is_opp": is_opp, "amount": amount,
            "opp_type": opp_type, "scenario": scenario, "is_convert": is_convert
        })
    wb.close()
    print(f"  [OK] 读取完成: {len(records)} 条记录 (跳过 {empty_dept} 条空部门)")
    return records


def calc_metrics(records):
    """计算一组记录的汇总指标。"""
    total = len(records)

    collab_cnt = sum(1 for r in records if r["collab"])
    dir_cnt = sum(1 for r in records if "总监" in r["collab"])
    ya_cnt = sum(1 for r in records if "云安" in r["collab"])

    opp_recs = [r for r in records if r["is_opp"] == "是"]
    proj_cnt = sum(1 for r in opp_recs if "项目" in r["opp_type"])
    std_cnt = sum(1 for r in opp_recs if "标品" in r["opp_type"])
    total_amt = sum(r["amount"] for r in opp_recs)

    conv_recs = [r for r in records if r["is_convert"] == "是"]
    conv_cnt = len(conv_recs)
    conv_amt = sum(r["amount"] for r in conv_recs)

    zero_breakers = set(r["manager"] for r in conv_recs if r["manager"])
    zero_breaker_cnt = len(zero_breakers)

    # 按转化记录统计五类场景
    scenarios = {"养龙虾": 0, "智能体": 0, "知识库": 0, "远航平台": 0, "其他": 0}
    for r in conv_recs:
        sc = (r["scenario"] or "").replace("​", "").replace("　", "").replace("\n", "").strip()
        if "养龙虾" in sc:
            scenarios["养龙虾"] += 1
        elif "智能体" in sc:
            scenarios["智能体"] += 1
        elif "知识库" in sc:
            scenarios["知识库"] += 1
        elif "远航" in sc or "远航平台" in sc:
            scenarios["远航平台"] += 1
        else:
            scenarios["其他"] += 1

    return {
        "走访数": total,
        "协同走访数": collab_cnt,
        "协同走访占比": collab_cnt / total if total else 0,
        "其中：总监": dir_cnt,
        "其中：云安": ya_cnt,
        "标品商机数": std_cnt,
        "项目商机数": proj_cnt,
        "总商机金额": total_amt,
        "转化数": conv_cnt,
        "转化金额": conv_amt,
        "破零人数": zero_breaker_cnt,
        **scenarios
    }


def safe_write(ws, row, col, value):
    """安全写入单元格，自动解除合并单元格。"""
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for mc in list(ws.merged_cells.ranges):
            if mc.min_row <= row <= mc.max_row and mc.min_col <= col <= mc.max_col:
                ws.unmerge_cells(str(mc))
                break
        cell = ws.cell(row, col)
    cell.value = value


def write_table(ws, depts, dept_metrics, total_metrics, core_keys,
                data_start_row, total_row, is_upper):
    """将一组部门的数据写入指定行范围。

    Args:
        ws:             worksheet 对象
        depts:          部门列表（按顺序）
        dept_metrics:   {dept: metrics_dict}，key 为部门名
        total_metrics:  本组合计的 metrics_dict
        core_keys:      核心指标键列表（12项，对应B-M列）
        data_start_row: 第一条数据所在行号
        total_row:      合计行行号
        is_upper:       True=上表（col8=项目商机数,整数），False=下表（col8=占比,百分比）
    """
    # ---- 1. 写入各部门数据行 ----
    for i, dept in enumerate(depts):
        row = data_start_row + i
        m = dept_metrics.get(dept, {})
        staff = STAFF_MAP.get(dept, 0)

        # A列：部门名称
        safe_write(ws, row, 1, dept)

        # B-M列（col 2-13）：核心指标
        for j, key in enumerate(core_keys):
            col = 2 + j
            if key == "AI破零率":
                val = m.get("破零人数", 0) / staff if staff else 0.0
            elif key == "标品商机占比":
                visits = m.get("走访数", 0)
                val = m.get("标品商机数", 0) / visits if visits else 0.0
            elif key == "破零人数":
                val = int(m.get("破零人数", 0))
            else:
                val = m.get(key, 0)
            safe_write(ws, row, col, val)

        # N、O列（col 14-15）：显式清空（旧模板可能有公式/值残留）
        safe_write(ws, row, 14, None)
        safe_write(ws, row, 15, None)

        # P-T列（col 16-20）：场景指标
        for j, key in enumerate(SCENARIO_KEYS):
            col = 16 + j
            safe_write(ws, row, col, m.get(key, 0))

    # ---- 2. 写入合计行 ----
    safe_write(ws, total_row, 1, "合计")
    total_staff = sum(STAFF_MAP.get(d, 0) for d in depts)

    for j, key in enumerate(core_keys):
        col = 2 + j
        if key == "AI破零率":
            val = total_metrics["破零人数"] / total_staff if total_staff else 0.0
        elif key == "标品商机占比":
            visits = total_metrics.get("走访数", 0)
            val = total_metrics.get("标品商机数", 0) / visits if visits else 0.0
        elif key == "破零人数":
            val = int(total_metrics.get("破零人数", 0))
        else:
            val = total_metrics.get(key, 0)
        safe_write(ws, total_row, col, val)

    # 合计行场景列
    for j, key in enumerate(SCENARIO_KEYS):
        col = 16 + j
        safe_write(ws, total_row, col, total_metrics.get(key, 0))

    # 合计行也清空N/O列
    safe_write(ws, total_row, 14, None)
    safe_write(ws, total_row, 15, None)

    # ---- 3. 数字格式设置 ----
    all_rows = list(range(data_start_row, data_start_row + len(depts))) + [total_row]
    for row in all_rows:
        ws.cell(row, 4).number_format = "0.0%"       # 协同走访占比
        ws.cell(row, 9).number_format = "#,##0"       # 总商机金额
        ws.cell(row, 11).number_format = "#,##0"      # 转化金额
        ws.cell(row, 12).number_format = "0"          # 破零人数
        ws.cell(row, 13).number_format = "0.0%"       # AI破零率
        if is_upper:
            ws.cell(row, 8).number_format = "0"       # 项目商机数（整数）
        else:
            ws.cell(row, 8).number_format = "0.0%"    # 标品商机占比（百分比）


def update_title_date(ws):
    """将第1行和第11行的标题日期更新为当天日期（0506-MMDD格式）。"""
    today = date.today()
    date_str = f"0506-{today.strftime('%m%d')}"
    new_title = f"AI标品专项工作开展情况（{date_str}）"

    safe_write(ws, 1, 1, new_title)
    safe_write(ws, 11, 1, new_title)
    print(f"  [date] 标题日期已更新: 0506-{today.strftime('%m%d')}")


def build_summary(all_depts, dept_metrics, total_metrics):
    """生成文字通报摘要。"""
    valid = [d for d in all_depts if d in dept_metrics and dept_metrics[d].get("走访数", 0) > 0]
    if not valid:
        return "【AI标品专项工作通报】\n暂无有效数据"

    best_visit = max(valid, key=lambda d: dept_metrics[d]["走访数"])
    best_std = max(valid, key=lambda d: dept_metrics[d]["标品商机数"])
    best_amt = max(valid, key=lambda d: dept_metrics[d]["总商机金额"])

    conv_depts = [d for d in valid if dept_metrics[d]["转化数"] > 0]
    not_conv = [d for d in valid if dept_metrics[d]["转化数"] == 0]

    return "\n".join([
        "【AI标品专项工作通报】",
        f"一、走访：累计{total_metrics['走访数']}家，协同{total_metrics['协同走访数']}"
        f"({total_metrics['协同走访占比']:.0%})，{best_visit}领先({dept_metrics[best_visit]['走访数']}家)",
        f"二、商机：标品{total_metrics['标品商机数']}个，项目{total_metrics['项目商机数']}个，"
        f"总金额{total_metrics['总商机金额']:,.0f}元，{best_std}标品领先，{best_amt}金额最高",
        f"三、转化：累计转化{total_metrics['转化数']}个({total_metrics['转化金额']:,.0f}元)，"
        f"破零{total_metrics['破零人数']}人，{'、'.join(conv_depts)}已破零"
        + (f"，{'、'.join(not_conv)}未破零" if not_conv else "")
    ])


def print_table(label, depts, dept_metrics):
    """控制台打印一张汇总表。"""
    header = (f"{'部门':<10} {'走访':>4} {'协同':>4} {'标品':>4} {'项目':>4} "
              f"{'金额':>10} {'转化':>4} {'破零':>4} {'破零率':>6} "
              f"{'养龙虾':>4} {'智能':>4} {'知识':>4} {'远航':>4} {'其他':>4}")
    print(f"\n  --- {label} ---")
    print("  " + "-" * (len(header) - 2))
    print("  " + header)
    print("  " + "-" * (len(header) - 2))
    for dept in depts:
        m = dept_metrics.get(dept, {})
        if not m:
            print(f"  {dept:<10} （无数据）")
            continue
        staff = STAFF_MAP.get(dept, 0)
        br = m["破零人数"] / staff if staff else 0
        print(f"  {dept:<10} {m['走访数']:>4} {m['协同走访数']:>4} "
              f"{m['标品商机数']:>4} {m['项目商机数']:>4} "
              f"{m['总商机金额']:>10,.0f} {m['转化数']:>4} "
              f"{m['破零人数']:>4} {br:>5.1%} "
              f"{m['养龙虾']:>4} {m['智能体']:>4} {m['知识库']:>4} "
              f"{m['远航平台']:>4} {m['其他']:>4}")
    print("  " + "-" * (len(header) - 2))


def build_narrative_report(all_dept_metrics, upper_total, lower_total, overall_total):
    """按模板格式生成文字版通报，写入 A20 单元格。

    Args:
        all_dept_metrics: {dept: metrics} 全量（含全部7个部门）
        upper_total:      BU单位（上表部门）合计指标
        lower_total:      政商单位（下表部门）合计指标
        overall_total:    全体合计指标
    Returns:
        文字通报字符串
    """
    today = date.today()
    date_str = f"0506-{today.strftime('%m%d')}"
    t = overall_total

    # ==================== 走访 ====================
    # BU
    bu_visits = upper_total["走访数"]
    bu_collab = upper_total["协同走访数"]
    bu_collab_pct = bu_collab / bu_visits if bu_visits else 0
    bu_dir = upper_total["其中：总监"]
    bu_ya = upper_total["其中：云安"]
    bu_best_visit = max(UPPER_DEPTS, key=lambda d: all_dept_metrics[d]["走访数"])

    # 政商
    zs_visits = lower_total["走访数"]
    zs_collab = lower_total["协同走访数"]
    zs_collab_pct = zs_collab / zs_visits if zs_visits else 0
    zs_dir = lower_total["其中：总监"]
    zs_ya = lower_total["其中：云安"]
    # 政商走访并列第一
    zs_visit_max_val = max(all_dept_metrics[d]["走访数"] for d in LOWER_DEPTS)
    zs_top_visits = [d for d in LOWER_DEPTS if all_dept_metrics[d]["走访数"] == zs_visit_max_val]

    # ==================== 商机 ====================
    # BU
    bu_std = upper_total["标品商机数"]
    bu_proj = upper_total["项目商机数"]
    bu_opp_total = bu_std + bu_proj
    bu_amt = upper_total["总商机金额"]
    bu_best_amt_dept = max(UPPER_DEPTS, key=lambda d: all_dept_metrics[d]["总商机金额"])
    bu_best_amt_val = all_dept_metrics[bu_best_amt_dept]["总商机金额"]
    bu_amt_pct = bu_best_amt_val / bu_amt if bu_amt else 0

    # 政商
    zs_std = lower_total["标品商机数"]
    zs_proj = lower_total["项目商机数"]
    zs_opp_total = zs_std + zs_proj
    zs_amt = lower_total["总商机金额"]
    zs_best_amt_dept = max(LOWER_DEPTS, key=lambda d: all_dept_metrics[d]["总商机金额"])
    zs_best_amt_val = all_dept_metrics[zs_best_amt_dept]["总商机金额"]

    # ==================== 转化 ====================
    # BU
    bu_conv = upper_total["转化数"]
    bu_conv_amt = upper_total["转化金额"]
    bu_breakers = upper_total["破零人数"]
    bu_total_staff = sum(STAFF_MAP[d] for d in UPPER_DEPTS)
    bu_breaker_rate = bu_breakers / bu_total_staff if bu_total_staff else 0
    bu_conv_depts = [d for d in UPPER_DEPTS if all_dept_metrics[d]["转化数"] > 0]
    bu_not_conv = [d for d in UPPER_DEPTS if all_dept_metrics[d]["转化数"] == 0]

    # 政商
    zs_conv = lower_total["转化数"]
    zs_conv_amt = lower_total["转化金额"]
    zs_breakers = lower_total["破零人数"]
    zs_total_staff = sum(STAFF_MAP[d] for d in LOWER_DEPTS)
    zs_breaker_rate = zs_breakers / zs_total_staff if zs_total_staff else 0
    zs_not_all = [d for d in LOWER_DEPTS if all_dept_metrics[d]["破零人数"] < STAFF_MAP[d]]

    # 整体
    total_conv = t["转化数"]
    total_conv_amt = t["转化金额"]
    total_breakers = t["破零人数"]
    total_staff = bu_total_staff + zs_total_staff
    overall_rate = total_breakers / total_staff if total_staff else 0

    # ==================== 场景分布 ====================
    scene_parts = []
    for key, label in [("知识库", "知识库"), ("养龙虾", "养虾"), ("智能体", "智能体"), ("远航平台", "远航平台")]:
        val = t.get(key, 0)
        if val > 0:
            scene_parts.append(f"{label}{val}个")
    scene_str = "、".join(scene_parts) if scene_parts else "暂无"

    # 未突破场景
    all_scenes = {"知识库", "养龙虾", "智能体", "远航平台"}
    hit_scenes = {k for k in all_scenes if t.get(k, 0) > 0}
    miss_scenes = all_scenes - hit_scenes
    miss_str = "、".join([{"养龙虾": "养虾", "知识库": "知识库", "智能体": "智能体", "远航平台": "远航平台"}.get(s, s) for s in sorted(miss_scenes)])

    # ==================== 组装 ====================
    lines = [
        f"【AI 标品专项工作进展通报（{date_str}）】",
        "一、走访情况",
        f"本期累计走访 {t['走访数']} 家（BU 单位 {bu_visits} 家，政商单位 {zs_visits} 家），"
        f"协同走访合计 {t['协同走访数']} 家，整体协同占比{t['协同走访占比']:.1%}。",
        f"BU 单位：走访 {bu_visits} 家，协同走访 {bu_collab} 家，协同占比 {bu_collab_pct:.0%}。"
        f"其中总监协同 {bu_dir} 家，云安协同 {bu_ya} 家。"
        f"{bu_best_visit}走访量最高（{all_dept_metrics[bu_best_visit]['走访数']} 家）。",
        f"政商单位：走访 {zs_visits} 家，协同走访 {zs_collab} 家，协同占比 {zs_collab_pct:.1%}。"
        f"其中总监协同 {zs_dir} 家，云安协同 {zs_ya} 家。"
        + (f"{'与'.join(zs_top_visits)}走访量并列第一（均为 {zs_visit_max_val} 家）。"
           if len(zs_top_visits) > 1 else
           f"{zs_top_visits[0]}走访量最高（{zs_visit_max_val} 家）。"),
        "二、商机情况",
        f"累计挖掘 AI 商机 {bu_opp_total + zs_opp_total} 个（标品 {t['标品商机数']} 个、项目 {t['项目商机数']} 个），"
        f"预估年化总金额约 {t['总商机金额']/10000:.1f} 万元。",
        f"BU 单位：挖掘商机 {bu_opp_total} 个（标品 {bu_std} 个、项目 {bu_proj} 个），"
        f"总金额约 {bu_amt/10000:.1f} 万元 / 年。"
        f"其中{bu_best_amt_dept}板块贡献突出，达 {bu_best_amt_val/10000:.1f} 万元 / 年，"
        f"占 BU 总额的 {bu_amt_pct:.1%}。",
        f"政商单位：挖掘商机 {zs_opp_total} 个（标品 {zs_std} 个、项目 {zs_proj} 个），"
        f"总金额约 {zs_amt/10000:.2f} 万元 / 年。"
        f"其中{zs_best_amt_dept}贡献最大，达 {zs_best_amt_val/10000:.2f} 万元 / 年。",
        "三、转化落地情况",
        f"本期累计实现商机转化 {total_conv} 个，转化金额 {total_conv_amt:,.0f} 元 / 年，"
        f"整体 AI 破零率 {overall_rate:.1%}。",
        f"BU 单位：转化 {bu_conv} 个，金额 {bu_conv_amt:,.0f} 元 / 年，破零 {bu_breakers} 人，"
        f"破零率 {bu_breaker_rate:.1%}。"
        + (f"仅{'、'.join(bu_conv_depts)}实现破零（转化场景为"
           f"\"{get_dept_scenes_str(all_dept_metrics, bu_conv_depts[0])}\"）"
           if len(bu_conv_depts) == 1 else "")
        + ("；" if bu_not_conv else "")
        + (f"{'、'.join(bu_not_conv)}尚未破零。" if bu_not_conv else ""),
        f"政商单位：转化 {zs_conv} 个，金额 {zs_conv_amt:,.0f} 元 / 年，破零 {zs_breakers} 人，"
        f"破零率 {zs_breaker_rate:.1%}。"
        + (f"{'、'.join(zs_not_all)}暂未全部破零。" if zs_not_all else "各单位已全部破零。"),
        f"从转化场景分布看，{scene_str}；{miss_str}场景暂待突破。",
    ]
    return "\n".join(lines)


def get_dept_scenes_str(dept_metrics, dept_name):
    """获取某部门转化记录中的场景名称字符串。"""
    # 需要从 records 获取，这里用已有的 metrics 推断
    # 实际上场景统计已经在 metrics 里了，取有值的场景
    m = dept_metrics[dept_name]
    scenes = []
    for key, label in [("知识库", "知识库"), ("养龙虾", "养龙虾"), ("智能体", "智能体"), ("远航平台", "远航平台")]:
        if m.get(key, 0) > 0:
            scenes.append(label)
    return "、".join(scenes) if scenes else "其他"


def fill_tongbao(filepath, groups, all_dept_metrics):
    """打开通报文件，写入两个表格的数据。

    Args:
        filepath:         AI通报.xlsx 路径
        groups:           {"upper": records_of_upper, "lower": records_of_lower}
        all_dept_metrics: {dept: metrics} 全量
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[SHEET_NAME]

    # 1. 更新标题日期
    update_title_date(ws)

    # 2. 上表：数字政务、数字企业、交通物流、教育 → 数据行4-7，合计行8
    upper_records = groups["upper"]
    upper_total = calc_metrics(upper_records)
    print(f"\n  [上表] 写入 {len(UPPER_DEPTS)} 个部门 → 行4-7，合计行8")
    write_table(ws, UPPER_DEPTS, all_dept_metrics, upper_total,
                UPPER_CORE_KEYS, data_start_row=4, total_row=8, is_upper=True)

    # 3. 下表：新塘政商、开发区政商、荔城政商 → 数据行14-16，合计行17
    lower_records = groups["lower"]
    lower_total = calc_metrics(lower_records)
    print(f"  [下表] 写入 {len(LOWER_DEPTS)} 个部门 → 行14-16，合计行17")
    write_table(ws, LOWER_DEPTS, all_dept_metrics, lower_total,
                LOWER_CORE_KEYS, data_start_row=14, total_row=17, is_upper=False)

    # 4. 文字版通报 → A20
    overall_total = calc_metrics(groups["upper"] + groups["lower"])
    narrative = build_narrative_report(all_dept_metrics, upper_total, lower_total, overall_total)
    safe_write(ws, 20, 1, narrative)

    wb.save(filepath)
    wb.close()
    print("  [OK] 写入完成")
    # 控制台也输出文字通报
    print("\n--- 文字版通报 (A20) ---")
    for line in narrative.split("\n"):
        print(f"  {line}")
    print("--- 通报结束 ---")


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------
def main():
    for f in [PAIDAN_FILE, TONGBAO_FILE]:
        if not os.path.exists(f):
            print(f"[ERR] 错误：找不到文件 {f}")
            sys.exit(1)

    print("=" * 60)
    print("  AI标品通报数据填充工具 v3.0 (通报模板v2 双表)")
    print("=" * 60)

    # ---- 第1步：读取 ----
    print("\n[1/4] 读取派单数据...")
    records = read_paidan(PAIDAN_FILE)
    if not records:
        print("[ERR] 无有效数据")
        sys.exit(1)

    # ---- 第2步：按部门计算指标 ----
    print("\n[2/4] 按部门计算指标...")
    groups = defaultdict(list)
    for r in records:
        groups[r["dept"]].append(r)
    all_dept_metrics = {d: calc_metrics(groups[d]) for d in groups}

    # 拆分上下表记录
    upper_records = [r for r in records if r["dept"] in UPPER_DEPTS]
    lower_records = [r for r in records if r["dept"] in LOWER_DEPTS]

    # 控制台输出
    print_table("上表 (通报模板v2 行4-7，合计行8)", UPPER_DEPTS, all_dept_metrics)
    print_table("下表 (通报模板v2 行14-16，合计行17)", LOWER_DEPTS, all_dept_metrics)

    # ---- 第3步：生成摘要 ----
    print("\n[3/4] 生成摘要...")
    all_depts = UPPER_DEPTS + LOWER_DEPTS
    summary = build_summary(all_depts, all_dept_metrics, calc_metrics(records))
    for line in summary.split("\n"):
        print(f"  [摘要] {line}")

    # ---- 第4步：写入 ----
    print("\n[4/4] 写入通报模板v2...")
    fill_tongbao(
        TONGBAO_FILE,
        {"upper": upper_records, "lower": lower_records},
        all_dept_metrics
    )

    print("\n" + "=" * 60)
    print("  [OK] 全部完成！AI通报.xlsx 的 通报模板v2 已更新。")
    print("=" * 60)


if __name__ == "__main__":
    main()
